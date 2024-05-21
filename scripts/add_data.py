from openpyxl import load_workbook
from django.contrib.auth.models import User
from applications.alumniprofile.models import Profile, Batch
from django.db import transaction
from AlumniConnect.views  import reg_no_gen

def add_data() :

    location = "scripts/data_acc.xlsx"
    wb = load_workbook(location)
    sheet = wb.active

    print("Adding data")
    with transaction.atomic():
        try:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                roll_no, name, email, sex, dob, programme, branch, batch_year,year_of_admission = row

                roll_no = str(int(roll_no))
                email = str(email)
                first_name, last_name = name.split(' ', 1)


                # User creation logic
                user, user_created = User.objects.get_or_create(
                    username=roll_no, 
                    first_name=first_name,
                    last_name=last_name,
                    defaults={'email': email})
                if user_created:
                    user.set_password(roll_no)
                    user.save()

                # Batch creation logic
                batch, created = Batch.objects.get_or_create(batch=int(batch_year))

                # Profile creation logic
                profile, profile_created = Profile.objects.get_or_create(
                    user=user,
                    defaults={
                        'email': email,
                        'roll_no': roll_no,
                        'name': name,
                        'sex': sex,
                        'programme': programme,
                        'branch': branch,
                        'batch': batch,  # use the batch instance directly
                        'date_of_birth': dob,  # Assuming dob is already a datetime object
                        "year_of_admission" : year_of_admission
                    }
                )

                if profile_created:
                    profile.reg_no = reg_no_gen(programme, branch, batch_year)
                    profile.save()
                    print(profile.name, profile.year_of_admission, profile.reg_no)

                

        except Exception as e :
            print(f'An error occured: {e}')
            raise